#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从数据底稿 Excel 生成图表数据 JSON，供 HTML 报告中的 Chart.js 使用。

用法:
  python generate_charts.py --excel 数据底稿.xlsx --output charts_data.json

输出 JSON 结构:
{
  "companies": ["万科", "金茂"],
  "periods": ["2024", "2025"],
  "charts": {
    "revenue": {
      "title": "营业收入与归母净利润",
      "type": "bar",
      "labels": ["万科 2024", "万科 2025", "金茂 2024", "金茂 2025"],
      "datasets": [
        {"label": "营业收入(亿)", "data": [2435, 2180, ...], "backgroundColor": "..."},
        {"label": "归母净利润(亿)", "data": [120, 98, ...], "backgroundColor": "..."}
      ]
    },
    "margin": { "title": "毛利率趋势", "type": "line", ... },
    "leverage": { "title": "负债水平", "type": "bar", ... },
    "assetStructure": { "title": "资产结构", "type": "bar", ... },
    "cashflow": { "title": "现金流", "type": "bar", ... }
  }
}
"""

import argparse, json, os, sys, re
from collections import OrderedDict

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ── 调色板 ──
BLUE = "rgba(47, 84, 150, 0.85)"
BLUE_LIGHT = "rgba(47, 84, 150, 0.4)"
ORANGE = "rgba(237, 125, 49, 0.85)"
ORANGE_LIGHT = "rgba(237, 125, 49, 0.4)"
GREEN = "rgba(68, 158, 115, 0.85)"
RED = "rgba(220, 78, 78, 0.85)"
GRAY = "rgba(150, 150, 150, 0.85)"
PALETTE = [
    "rgba(47, 84, 150, 0.85)",     # 深蓝
    "rgba(237, 125, 49, 0.85)",    # 橙
    "rgba(68, 158, 115, 0.85)",    # 绿
    "rgba(220, 78, 78, 0.85)",     # 红
    "rgba(140, 100, 180, 0.85)",   # 紫
    "rgba(0, 170, 180, 0.85)",     # 青
]


def read_pivot(excel_path: str) -> dict:
    """读取汇总透视表 Sheet，返回结构化数据。

    假设透视表结构：第1列=指标名, 第2列=单位, 第3列起=数据列。
    表头第1行：指标 | 单位 | 公司A 期间1 | 公司A 期间2 | ...
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 第一个 sheet = 汇总透视表

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("透视表至少需要表头+1行数据")

    # 解析表头：第1列="指标", 第2列="单位", 第3列起=公司 期间
    headers = [str(h).strip() if h else "" for h in rows[0]]
    if headers[0] != "指标":
        # 兼容旧格式：第1行公司名，第2行年份
        if len(rows) >= 2:
            h1 = [str(h).strip() if h else "" for h in rows[0]]
            h2 = [str(h).strip() if h else "" for h in rows[1]]
            col_labels = []
            # 第1列是指标（可能跨2行），第2列可能是单位或直接数据
            data_start_row = 2
            for ci in range(len(h1)):
                if ci < 2:
                    continue  # 跳过指标名和单位列
                c1 = h1[ci] if h1[ci] else ""
                c2 = h2[ci] if ci < len(h2) and h2[ci] else ""
                label = f"{c1} {c2}".strip()
                col_labels.append(label)
            headers = ["指标", "单位"] + col_labels
        rows = rows[data_start_row:]
    else:
        rows = rows[1:]  # 跳过表头行

    # 解析数据
    indicators = OrderedDict()
    companies = OrderedDict()
    periods = []
    col_pairs = []  # [(公司, 期间)]

    for ci in range(2, len(headers)):
        label = headers[ci]
        parts = label.split()
        if len(parts) >= 2:
            co = parts[0]
            period = " ".join(parts[1:])
        else:
            co = label
            period = ""
        col_pairs.append((co, period))
        companies[co] = True
        if period and period not in periods:
            periods.append(period)

    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        unit = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        values = []
        for ci in range(2, min(len(row), len(headers))):
            val = row[ci]
            if val is None or str(val).strip() in ("-", "", "NA"):
                values.append(None)
            else:
                try:
                    values.append(float(str(val).replace(",", "").replace(" ", "")))
                except ValueError:
                    values.append(None)
        if name and any(v is not None for v in values):
            indicators[name] = {"unit": unit, "values": values}

    return {
        "companies": list(companies.keys()),
        "periods": periods,
        "col_pairs": [(co, pr) for co, pr in col_pairs],
        "indicators": indicators,
    }


def _get_indicator(data: dict, name: str) -> tuple:
    """获取指标的值列表和单位，找不到返回 ([], '')."""
    ind = data["indicators"].get(name)
    if ind:
        return ind["values"], ind["unit"]
    return [], ""


def _build_labels(data: dict) -> list:
    return [f"{co} {pr}" for co, pr in data["col_pairs"]]


def _color_for_company(co: str, companies: list) -> str:
    idx = companies.index(co) if co in companies else 0
    return PALETTE[idx % len(PALETTE)]


def build_charts(data: dict) -> OrderedDict:
    """从透视表数据构建所有图表的 config JSON。"""
    charts = OrderedDict()
    labels = _build_labels(data)
    companies = data["companies"]
    col_pairs = data["col_pairs"]

    # ── 1. 收入与利润 柱状图 ──
    rev_vals, rev_unit = _get_indicator(data, "营业收入")
    profit_vals, profit_unit = _get_indicator(data, "归母净利润")
    if rev_vals or profit_vals:
        datasets = []
        if rev_vals:
            datasets.append({
                "label": f"营业收入({rev_unit})",
                "data": rev_vals,
                "backgroundColor": BLUE,
                "borderColor": BLUE.replace("0.85", "1"),
                "borderWidth": 1,
            })
        if profit_vals:
            datasets.append({
                "label": f"归母净利润({profit_unit})",
                "data": profit_vals,
                "backgroundColor": ORANGE,
                "borderColor": ORANGE.replace("0.85", "1"),
                "borderWidth": 1,
            })
        if datasets:
            charts["revenue"] = {
                "title": "营业收入与归母净利润",
                "type": "bar",
                "labels": labels,
                "datasets": datasets,
                "options": {
                    "responsive": True,
                    "plugins": {"legend": {"position": "top"}},
                    "scales": {"y": {"beginAtZero": True, "title": {"display": True, "text": rev_unit or ""}}},
                },
            }

    # ── 2. 毛利率 折线图 ──
    margin_vals, _ = _get_indicator(data, "毛利率")
    if margin_vals:
        datasets = []
        for ci, (co, _) in enumerate(col_pairs):
            pass
        # 每条公司一条线
        for co_idx, co in enumerate(companies):
            co_vals = []
            co_labels = []
            for ci, (c, pr) in enumerate(col_pairs):
                if c == co:
                    co_labels.append(pr)
                    co_vals.append(margin_vals[ci] if ci < len(margin_vals) else None)
            if co_vals:
                datasets.append({
                    "label": co,
                    "data": co_vals,
                    "borderColor": _color_for_company(co, companies).replace("0.85", "1"),
                    "backgroundColor": _color_for_company(co, companies),
                    "tension": 0.3,
                    "fill": False,
                })
        if datasets:
            charts["margin"] = {
                "title": "毛利率趋势",
                "type": "line",
                "labels": data["periods"],
                "datasets": datasets,
                "options": {
                    "responsive": True,
                    "plugins": {"legend": {"position": "top"}},
                    "scales": {"y": {"beginAtZero": False, "title": {"display": True, "text": "%"}}},
                },
            }

    # ── 3. 负债水平 柱状图 ──
    debt_ratio, _ = _get_indicator(data, "资产负债率")
    net_debt_ratio, _ = _get_indicator(data, "净负债率")
    if debt_ratio or net_debt_ratio:
        datasets = []
        if debt_ratio:
            datasets.append({
                "label": "资产负债率(%)",
                "data": debt_ratio,
                "backgroundColor": RED,
                "borderColor": RED.replace("0.85", "1"),
                "borderWidth": 1,
            })
        if net_debt_ratio:
            datasets.append({
                "label": "净负债率(%)",
                "data": net_debt_ratio,
                "backgroundColor": GRAY,
                "borderColor": GRAY.replace("0.85", "1"),
                "borderWidth": 1,
            })
        if datasets:
            charts["leverage"] = {
                "title": "负债水平",
                "type": "bar",
                "labels": labels,
                "datasets": datasets,
                "options": {
                    "responsive": True,
                    "plugins": {"legend": {"position": "top"}},
                    "scales": {"y": {"beginAtZero": True, "title": {"display": True, "text": "%"}}},
                },
            }

    # ── 4. 资产结构 柱状图 ──
    total_assets, unit_a = _get_indicator(data, "总资产")
    total_debt, unit_d = _get_indicator(data, "总负债")
    invest_prop, _ = _get_indicator(data, "投资性房地产")
    inventory, _ = _get_indicator(data, "存货总额")
    datasets = []
    if total_assets:
        datasets.append({
            "label": f"总资产({unit_a})",
            "data": total_assets,
            "backgroundColor": BLUE,
            "borderColor": BLUE.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if total_debt:
        datasets.append({
            "label": f"总负债({unit_d})",
            "data": total_debt,
            "backgroundColor": RED,
            "borderColor": RED.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if invest_prop:
        datasets.append({
            "label": f"投资性房地产({unit_a})",
            "data": invest_prop,
            "backgroundColor": GREEN,
            "borderColor": GREEN.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if inventory:
        datasets.append({
            "label": f"存货({unit_a})",
            "data": inventory,
            "backgroundColor": ORANGE,
            "borderColor": ORANGE.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if datasets:
        charts["assetStructure"] = {
            "title": "资产与负债结构",
            "type": "bar",
            "labels": labels,
            "datasets": datasets,
            "options": {
                "responsive": True,
                "plugins": {"legend": {"position": "top"}},
                "scales": {"y": {"beginAtZero": True, "title": {"display": True, "text": unit_a or ""}}},
            },
        }

    # ── 5. 现金流 柱状图 ──
    op_cf, cf_unit = _get_indicator(data, "经营活动现金流净额")
    inv_cf, _ = _get_indicator(data, "投资活动现金流净额")
    fin_cf, _ = _get_indicator(data, "筹资活动现金流净额")
    datasets = []
    if op_cf:
        datasets.append({
            "label": f"经营活动({cf_unit})",
            "data": op_cf,
            "backgroundColor": GREEN,
            "borderColor": GREEN.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if inv_cf:
        datasets.append({
            "label": f"投资活动({cf_unit})",
            "data": inv_cf,
            "backgroundColor": ORANGE,
            "borderColor": ORANGE.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if fin_cf:
        datasets.append({
            "label": f"筹资活动({cf_unit})",
            "data": fin_cf,
            "backgroundColor": GRAY,
            "borderColor": GRAY.replace("0.85", "1"),
            "borderWidth": 1,
        })
    if datasets:
        charts["cashflow"] = {
            "title": "现金流结构",
            "type": "bar",
            "labels": labels,
            "datasets": datasets,
            "options": {
                "responsive": True,
                "plugins": {"legend": {"position": "top"}},
                "scales": {"y": {"title": {"display": True, "text": cf_unit or ""}}},
            },
        }

    return charts


def main():
    parser = argparse.ArgumentParser(description="从数据底稿生成图表数据JSON")
    parser.add_argument("--excel", required=True, help="数据底稿.xlsx 路径")
    parser.add_argument("--output", required=True, help="输出 JSON 路径")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"错误: Excel不存在: {args.excel}", file=sys.stderr)
        sys.exit(1)

    data = read_pivot(args.excel)
    charts = build_charts(data)

    output = {
        "companies": data["companies"],
        "periods": data["periods"],
        "charts": charts,
    }

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"图表数据已生成: {args.output}")
    print(f"  公司: {', '.join(data['companies'])}")
    print(f"  期间: {', '.join(data['periods'])}")
    print(f"  图表: {', '.join(charts.keys())}")


if __name__ == "__main__":
    main()
