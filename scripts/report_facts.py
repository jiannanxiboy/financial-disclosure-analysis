#!/usr/bin/env python3
"""Build a cell-addressed fact package from 数据底稿.xlsx and verify report claims."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from _common import atomic_write_json


def extract_facts(workbook_path: str | Path) -> dict:
    path = Path(workbook_path).resolve()
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "汇总透视表" not in workbook.sheetnames:
        raise ValueError("工作簿缺少 汇总透视表")
    sheet = workbook["汇总透视表"]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    columns = []
    for col, label in enumerate(headers[2:], start=3):
        text = str(label or "").strip()
        if " " not in text:
            raise ValueError(f"汇总列无法拆分公司与期间: {text!r}")
        company, period = text.rsplit(" ", 1)
        columns.append((col, company, period))
    facts = []
    for row in range(2, sheet.max_row + 1):
        indicator = str(sheet.cell(row, 1).value or "").strip()
        unit = str(sheet.cell(row, 2).value or "").strip()
        if not indicator:
            continue
        for col, company, period in columns:
            value = sheet.cell(row, col).value
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                continue
            facts.append({
                "indicator": indicator,
                "company": company,
                "period": period,
                "value": float(value),
                "unit": unit,
                "cell": f"汇总透视表!{get_column_letter(col)}{row}",
            })
    workbook.close()
    return {"schema_version": 1, "workbook": str(path), "facts": facts}


def _compare(actual: float, operator: str, expected: float, tolerance: float) -> bool:
    if operator == "eq":
        return math.isclose(actual, expected, rel_tol=tolerance, abs_tol=tolerance)
    if operator == "lt":
        return actual < expected
    if operator == "lte":
        return actual <= expected
    if operator == "gt":
        return actual > expected
    if operator == "gte":
        return actual >= expected
    raise ValueError(f"不支持的比较操作: {operator}")


def verify_claims(fact_package: dict, claims: list[dict]) -> list[dict]:
    index = {
        (fact["indicator"], fact["company"], fact["period"]): fact
        for fact in fact_package["facts"]
    }
    results = []
    for claim in claims:
        check_results = []
        for check in claim.get("checks", []):
            operator = check.get("operator", "eq")
            indicator = check["indicator"]
            evidence = []
            passed = False
            message = ""
            if operator in {"eq", "lt", "lte", "gt", "gte"}:
                key = (indicator, check["company"], str(check["period"]))
                fact = index.get(key)
                if fact:
                    evidence.append(fact)
                    passed = _compare(
                        fact["value"], operator, float(check["expected"]),
                        float(check.get("tolerance", 1e-6)),
                    )
                message = f"{key} {operator} {check['expected']}"
            elif operator in {"all_lt", "all_gt"}:
                left_period = str(check["left_period"])
                right_period = str(check["right_period"])
                companies = check.get("companies") or sorted({
                    fact["company"] for fact in fact_package["facts"] if fact["indicator"] == indicator
                })
                comparisons = []
                for company in companies:
                    left = index.get((indicator, company, left_period))
                    right = index.get((indicator, company, right_period))
                    if left and right:
                        evidence.extend([left, right])
                        comparisons.append(
                            left["value"] < right["value"] if operator == "all_lt"
                            else left["value"] > right["value"]
                        )
                    else:
                        comparisons.append(False)
                passed = bool(comparisons) and all(comparisons)
                message = f"{indicator}: {left_period} {operator} {right_period} for {len(companies)} companies"
            elif operator in {"max_company", "min_company"}:
                period = str(check["period"])
                candidates = [
                    fact for fact in fact_package["facts"]
                    if fact["indicator"] == indicator and fact["period"] == period
                ]
                evidence.extend(candidates)
                if candidates:
                    selected = (max if operator == "max_company" else min)(candidates, key=lambda item: item["value"])
                    passed = selected["company"] == check["expected"]
                    message = f"{operator}={selected['company']}, expected={check['expected']}"
            else:
                raise ValueError(f"不支持的 claim operator: {operator}")
            check_results.append({
                "operator": operator,
                "passed": passed,
                "message": message,
                "evidence": evidence,
            })
        verified = bool(check_results) and all(item["passed"] for item in check_results)
        results.append({
            "id": claim["id"],
            "text": claim["text"],
            "status": "verified" if verified else "failed",
            "checks": check_results,
        })
    return results


def build_package(workbook: str | Path, claims_path: str | Path | None = None) -> dict:
    package = extract_facts(workbook)
    claims = []
    if claims_path:
        raw = json.loads(Path(claims_path).read_text(encoding="utf-8"))
        claims = raw.get("claims", []) if isinstance(raw, dict) else raw if isinstance(raw, list) else []
    package["claims"] = verify_claims(package, claims) if claims else []
    package["status"] = (
        "failed" if any(claim["status"] != "verified" for claim in package["claims"])
        else "verified" if package["claims"] else "facts_only"
    )
    return package


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--claims", help="JSON claim definitions")
    parser.add_argument("--output", required=True)
    parser.add_argument("--allow-unverified", action="store_true")
    args = parser.parse_args()
    package = build_package(args.workbook, args.claims)
    atomic_write_json(args.output, package)
    print(json.dumps({
        "output": str(Path(args.output).resolve()),
        "facts": len(package["facts"]),
        "claims": len(package["claims"]),
        "status": package["status"],
    }, ensure_ascii=False))
    return 0 if args.allow_unverified or package["status"] != "failed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
