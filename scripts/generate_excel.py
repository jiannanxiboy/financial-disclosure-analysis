#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从TSV文件生成Excel分析报告。

用法:
  python generate_excel.py --config <config.json>

config.json 格式:
{
  "companies": [
    {
      "name": "万科",
      "periods": {
        "2022": "path/to/万科_2022.tsv",
        "2023": "path/to/万科_2023.tsv",
        "2024": "path/to/万科_2024.tsv",
        "202403": "path/to/万科_202403.tsv"
      }
    }
  ],
  "indicators": ["营业收入", "营业成本", ...],
  "output": "path/to/output.xlsx"
}

periods 的 key 为期间标签（如年度"2024"、季度"202403"、"2024H1"等），
作为列标题和Sheet名使用。
兼容旧字段名 years。

TSV格式（每文件一个公司+期间组合）:
  指标名称 \t 数据 \t 备注
"""

import argparse, csv, json, os, sys, re
from collections import OrderedDict
from pathlib import Path

from _common import atomic_write_json
from financial_validation import MetricRecord, validate_records

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ── 样式定义 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# 指标行背景色区分：偶数行浅蓝
STRIPE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

HEADER_ALIASES = {
    "indicator": {"指标名称", "名称", "指标"},
    "value": {"数据", "金额/比例", "值"},
    "note": {"备注", "来源备注", "来源"},
}

MISSING_VALUES = {"", "-", "--", "—", "NA", "N/A", "NULL", "NIL", "不适用", "未披露"}

DEFAULT_INDICATOR_DEFINITIONS = {
    **{
        name: {
            "canonical_unit": "亿",
            "accepted_units": {
                "亿": 1,
                "亿元": 1,
                "百万元": 0.01,
                "万元": 0.0001,
                "元": 0.00000001,
            },
        }
        for name in (
            "合约销售额", "营业收入", "总资产", "归母权益", "毛利", "归母净利润",
            "核心净利润", "总负债", "有息负债(借贷总额)", "现金及现金等价物",
            "经营活动现金流净额", "融资活动现金流净额",
        )
    },
    **{
        name: {
            "canonical_unit": "万平方米",
            "accepted_units": {
                "万平方米": 1,
                "万平米": 1,
                "平方米": 0.0001,
                "平米": 0.0001,
            },
        }
        for name in (
            "合约销售面积", "土地储备面积(总计)", "土地储备面积(应占)",
            "结算面积", "新增土储面积",
        )
    },
    **{
        name: {
            "canonical_unit": "%",
            "accepted_units": {"%": 1, "％": 1},
        }
        for name in (
            "毛利率", "归母净利率", "核心净利率", "加权平均净资产收益率",
            "资产负债率", "净负债率", "短期借贷占有息负债比", "加权平均融资成本",
        )
    },
    "平均售价": {
        "canonical_unit": "万元/平方米",
        "accepted_units": {
            "万元/平方米": 1,
            "万元/平米": 1,
            "万/平方米": 1,
            "万/平米": 1,
            "元/平方米": 0.0001,
            "元/平米": 0.0001,
        },
    },
    "每股基本盈利": {
        "canonical_unit": "元",
        "accepted_units": {"元": 1, "人民币元": 1, "分": 0.01},
    },
}


def parse_value(raw: str) -> tuple:
    """解析披露值，保留尾部单位；支持括号负数和常见缺失标记。"""
    if raw is None or raw.strip().upper() in MISSING_VALUES:
        return (None, "")
    s = raw.strip().replace(",", "").replace("，", "").replace(" ", "").replace("　", "")
    negative_match = re.fullmatch(r"\(([+-]?(?:\d+(?:\.\d*)?|\.\d+))\)(.*)", s)
    negative = negative_match is not None
    if negative_match:
        s = negative_match.group(1) + negative_match.group(2)
    s = re.sub(r"^(人民币|RMB|CNY|HKD|HK\$|￥|¥)", "", s, flags=re.IGNORECASE)
    m = re.fullmatch(r'([+-]?(?:\d+(?:\.\d*)?|\.\d+))(.*)', s)
    if m:
        try:
            number = float(m.group(1))
            return (-number if negative else number, m.group(2).strip())
        except ValueError:
            return (None, s)
    return (None, s)


def load_indicator_definitions(config: dict) -> dict:
    """合并内置单位规则与配置覆盖，返回可直接用于换算的规则。"""
    definitions = {
        name: {
            "canonical_unit": rule["canonical_unit"],
            "accepted_units": dict(rule["accepted_units"]),
        }
        for name, rule in DEFAULT_INDICATOR_DEFINITIONS.items()
    }
    for name, rule in config.get("indicator_definitions", {}).items():
        canonical = str(rule.get("canonical_unit", "")).strip()
        accepted = rule.get("accepted_units", {})
        if not canonical or not isinstance(accepted, dict):
            raise ValueError(f"指标 {name} 的单位定义必须包含 canonical_unit 和 accepted_units")
        try:
            conversions = {str(unit).strip(): float(factor) for unit, factor in accepted.items()}
        except (TypeError, ValueError) as exc:
            raise ValueError(f"指标 {name} 的单位换算系数必须是数字") from exc
        conversions.setdefault(canonical, 1.0)
        definitions[name] = {"canonical_unit": canonical, "accepted_units": conversions}
    return definitions


def normalize_value(indicator: str, raw: str, definitions: dict) -> tuple:
    """把单点数据换算到指标标准单位；未知单位立即报错，避免静默混用。"""
    number, raw_unit = parse_value(raw)
    rule = definitions.get(indicator)
    if number is None:
        return None, rule["canonical_unit"] if rule else ""
    if not rule:
        return number, raw_unit
    canonical = rule["canonical_unit"]
    if not raw_unit:
        return number, canonical
    factors = rule["accepted_units"]
    if raw_unit not in factors:
        raise ValueError(
            f"指标 {indicator} 的单位 {raw_unit!r} 无法换算为 {canonical!r}；"
            "请在 indicator_definitions 中补充规则"
        )
    return number * factors[raw_unit], canonical


def style_header(ws, row, col_count):
    """给表头行加样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_row(ws, row, col_count, is_stripe: bool):
    """给数据行加样式"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if c == 1:
            cell.alignment = DATA_ALIGN_LEFT
        else:
            cell.alignment = DATA_ALIGN_CENTER
        if is_stripe:
            cell.fill = STRIPE_FILL


def auto_width(ws, col_count, min_width=10, max_width=36):
    """自适应列宽"""
    for c in range(1, col_count + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            for val in row:
                if val:
                    # 中文字符大约占2个英文宽度
                    length = 0
                    for ch in str(val):
                        length += 2 if '一' <= ch <= '鿿' else 1
                    max_len = max(max_len, length)
        width = max(min_width, min(max_len + 4, max_width))
        ws.column_dimensions[col_letter].width = width


def read_tsv(path: str) -> OrderedDict:
    """读取TSV，返回 {指标名: (金额/比例, 备注)}"""
    data = OrderedDict()
    if not os.path.exists(path):
        print(f"  [WARN] TSV不存在: {path}")
        return data
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        rows = list(reader)
    if not rows:
        return data

    header = [item.strip() for item in rows[0]]
    indexes = {}
    for canonical, aliases in HEADER_ALIASES.items():
        matches = [idx for idx, value in enumerate(header) if value in aliases]
        if len(matches) > 1:
            raise ValueError(f"TSV 表头字段重复: {canonical} ({path})")
        if matches:
            indexes[canonical] = matches[0]
    missing = [name for name in ("indicator", "value") if name not in indexes]
    if missing:
        raise ValueError(f"TSV 表头缺少必要字段 {missing}: {path}; 实际表头={header}")

    for line_no, row in enumerate(rows[1:], start=2):
        if not row or not any(item.strip() for item in row):
            continue
        name = row[indexes["indicator"]].strip() if indexes["indicator"] < len(row) else ""
        val = row[indexes["value"]].strip() if indexes["value"] < len(row) else ""
        note_idx = indexes.get("note")
        note = row[note_idx].strip() if note_idx is not None and note_idx < len(row) else ""
        if not name:
            raise ValueError(f"TSV 第 {line_no} 行缺少指标名称: {path}")
        if name in data:
            raise ValueError(f"TSV 第 {line_no} 行指标重复: {name} ({path})")
        data[name] = (val, note)
    return data


def build_pivot(all_data: dict, indicator_order: list, indicator_definitions: dict | None = None) -> tuple:
    """
    all_data: {(公司, 年份): {指标: (值, 备注)}}
    indicator_order: 指标名列表（有序）
    返回: (columns, rows) 用于写入透视表
      columns: ["指标", "单位", "万科 2024", "万科 2025", ...]
      rows: [["营业收入", "亿", 2435, 2180, ...], ...]
    数据单元格为纯数字，可求和。
    """
    # 收集所有公司-期间组合，按config顺序
    col_labels = []
    for (co, period) in all_data.keys():
        label = f"{co} {period}"
        if label not in col_labels:
            col_labels.append(label)

    definitions = indicator_definitions or load_indicator_definitions({})
    rows = []
    for ind in indicator_order:
        # 第一遍：收集所有值并推断单位
        raw_vals = []  # [(公司期间label, 数值, 单位)]
        for (co, period) in all_data.keys():
            label = f"{co} {period}"
            if label not in col_labels:
                continue
            vals = all_data.get((co, period), {})
            if ind in vals:
                num, unit = normalize_value(ind, vals[ind][0], definitions)
                raw_vals.append((label, num, unit))
            else:
                raw_vals.append((label, None, ""))

        # 有定义时使用标准单位；无定义指标要求原始单位保持一致。
        unit = definitions.get(ind, {}).get("canonical_unit", "")
        observed_units = {u for _, num, u in raw_vals if num is not None and u}
        if not unit and len(observed_units) > 1:
            raise ValueError(f"指标 {ind} 存在多个单位 {sorted(observed_units)}，但没有单位换算规则")
        if not unit and observed_units:
            unit = next(iter(observed_units))

        # 至少有一列有数据才保留
        if all(num is None for _, num, _ in raw_vals):
            continue

        row = [ind, unit]
        for _, num, _ in raw_vals:
            if num is not None:
                row.append(num)
            else:
                row.append("-")
        rows.append(row)

    return ["指标", "单位"] + col_labels, rows


def build_metric_records(all_data: dict, indicator_definitions: dict, source_paths: dict) -> list[MetricRecord]:
    records: list[MetricRecord] = []
    for (company, period), data in all_data.items():
        for indicator, (raw_value, note) in data.items():
            raw_number, raw_unit = parse_value(raw_value)
            normalized_value, normalized_unit = normalize_value(indicator, raw_value, indicator_definitions)
            if normalized_value is None:
                status = "missing"
            elif raw_unit and normalized_unit and raw_unit != normalized_unit:
                status = "converted"
            elif indicator in indicator_definitions:
                status = "normalized"
            else:
                status = "raw"
            records.append(MetricRecord(
                company=company,
                period=str(period),
                indicator=indicator,
                raw_value=raw_value,
                raw_unit=raw_unit,
                normalized_value=normalized_value,
                normalized_unit=normalized_unit,
                source_file=source_paths.get((company, period), ""),
                source_location=note,
                status=status,
            ))
    return records


def write_pivot_sheet(wb, columns, rows):
    """写入汇总透视表"""
    ws = wb.active
    ws.title = "汇总透视表"

    # 写表头
    for ci, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col_name)
    style_header(ws, 1, len(columns))

    # 写数据
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri + 2, column=ci, value=val)
        style_data_row(ws, ri + 2, len(columns), ri % 2 == 1)

    # 冻结首行 + 前两列（指标、单位）
    ws.freeze_panes = "C2"
    auto_width(ws, len(columns))


def write_validation_sheet(wb, summary: dict):
    ws = wb.create_sheet(title="校验摘要")
    headers = ["级别", "类别", "公司", "期间", "指标", "说明"]
    for ci, value in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=value)
    style_header(ws, 1, len(headers))
    issues = summary["errors"] + summary["warnings"] + summary["missing"] + summary["conversions"]
    for ri, issue in enumerate(issues, start=2):
        values = [
            issue["severity"], issue["category"], issue["company"], issue["period"],
            issue["indicator"], issue["message"],
        ]
        for ci, value in enumerate(values, 1):
            ws.cell(row=ri, column=ci, value=value)
        style_data_row(ws, ri, len(headers), ri % 2 == 0)
    if not issues:
        ws.cell(row=2, column=1, value="info")
        ws.cell(row=2, column=6, value="未发现异常、缺失或单位换算记录")
        style_data_row(ws, 2, len(headers), False)
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers), max_width=60)


def write_detail_sheets(wb, all_data: dict, records: list[MetricRecord] | None = None):
    """为每个公司-期间组合写明细Sheet"""
    record_map = {
        (record.company, record.period, record.indicator): record
        for record in (records or [])
    }
    for (co, period), data in all_data.items():
        sheet_name = f"{co}_{period}"[:31]  # Excel sheet名最长31字符
        if sheet_name in wb.sheetnames:
            # 已存在则追加后缀
            for i in range(2, 100):
                alt = f"{co}_{period}_{i}"[:31]
                if alt not in wb.sheetnames:
                    sheet_name = alt
                    break

        ws = wb.create_sheet(title=sheet_name)

        # 表头
        headers = ["指标名称", "原始值", "原始单位", "标准值", "标准单位", "来源备注", "状态"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        style_header(ws, 1, len(headers))

        # 数据
        for ri, (name, (val, note)) in enumerate(data.items()):
            record = record_map.get((co, str(period), name))
            values = [
                name,
                val,
                record.raw_unit if record else "",
                record.normalized_value if record else "",
                record.normalized_unit if record else "",
                note,
                record.status if record else "raw",
            ]
            for ci, value in enumerate(values, 1):
                ws.cell(row=ri + 2, column=ci, value=value)
            style_data_row(ws, ri + 2, len(headers), ri % 2 == 1)

        ws.freeze_panes = "A2"
        auto_width(ws, len(headers))
        # 备注列宽一些
        ws.column_dimensions["F"].width = 50


def main():
    parser = argparse.ArgumentParser(description="从TSV生成Excel分析报告")
    parser.add_argument("--config", required=True, help="JSON配置文件路径")
    parser.add_argument("--output", help="覆盖config中的Excel输出路径")
    parser.add_argument("--records-output", help="覆盖结构化记录JSON输出路径")
    parser.add_argument("--validation-output", help="覆盖校验摘要JSON输出路径")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser().resolve()
    config_dir = config_path.parent

    def resolve_config_path(value):
        path = Path(value).expanduser()
        return path if path.is_absolute() else config_dir / path

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    companies = cfg.get("companies", [])
    indicator_order = cfg.get("indicators", [])
    indicator_definitions = load_indicator_definitions(cfg)
    output = (
        Path(args.output).expanduser().resolve()
        if args.output else resolve_config_path(cfg.get("output", "output.xlsx"))
    )

    if not companies:
        print("错误: config中未指定companies")
        sys.exit(1)

    # ── 读取所有TSV ──
    all_data = OrderedDict()  # {(公司, 期间): {指标: (值, 备注)}}
    all_indicators = OrderedDict()
    source_paths = {}

    for comp in companies:
        name = comp["name"]
        periods = comp.get("periods", comp.get("years", {}))  # 兼容旧字段名
        for period, tsv_path in periods.items():
            resolved_tsv = resolve_config_path(tsv_path).resolve()
            print(f"读取: {name} {period} → {resolved_tsv}")
            data = read_tsv(resolved_tsv)
            all_data[(name, period)] = data
            source_paths[(name, period)] = str(resolved_tsv)
            for ind in data:
                all_indicators[ind] = True
            print(f"  提取 {len(data)} 个指标")

    # ── 确定指标顺序 ──
    if indicator_order:
        # 使用用户指定的顺序，未列出的追加到末尾
        final_order = [i for i in indicator_order if i in all_indicators]
        for i in all_indicators:
            if i not in final_order:
                final_order.append(i)
    else:
        final_order = list(all_indicators.keys())

    # ── 生成Excel ──
    wb = Workbook()

    records = build_metric_records(all_data, indicator_definitions, source_paths)
    validation_cfg = cfg.get("validation", {})
    validation_summary = validate_records(
        records,
        margin_tolerance=float(validation_cfg.get("margin_tolerance", 0.5)),
        net_margin_tolerance=float(validation_cfg.get("net_margin_tolerance", 0.2)),
        balance_tolerance_pct=float(validation_cfg.get("balance_tolerance_pct", 1.0)),
        yoy_warning_pct=float(validation_cfg.get("yoy_warning_pct", 100.0)),
    )

    # Sheet 1 — 汇总透视表
    columns, rows = build_pivot(all_data, final_order, indicator_definitions)
    write_pivot_sheet(wb, columns, rows)
    print(f"\n汇总透视表: {len(rows)} 行 × {len(columns)} 列")

    write_validation_sheet(wb, validation_summary)
    print(
        "校验摘要: "
        f"错误 {validation_summary['stats']['errors']} / "
        f"警告 {validation_summary['stats']['warnings']} / "
        f"缺失 {validation_summary['stats']['missing']}"
    )

    # Sheet 2+ — 各公司明细
    write_detail_sheets(wb, all_data, records)
    print(f"明细Sheet: {len(all_data)} 个")

    # 保存
    os.makedirs(output.parent, exist_ok=True)
    wb.save(output)
    output_path = output.resolve()
    records_output = (
        Path(args.records_output).expanduser().resolve()
        if args.records_output else resolve_config_path(
            cfg.get("records_output", output_path.with_name("normalized_records.json"))
        )
    )
    validation_output = (
        Path(args.validation_output).expanduser().resolve()
        if args.validation_output else resolve_config_path(
            cfg.get("validation_output", output_path.with_name("validation_summary.json"))
        )
    )
    atomic_write_json(records_output, {
        "schema_version": 1,
        "records": [record.to_dict() for record in records],
    })
    atomic_write_json(validation_output, validation_summary)
    print(f"\nExcel已生成: {output}")
    print(f"结构化记录: {records_output}")
    print(f"校验结果: {validation_output}")
    if validation_summary["errors"]:
        print("错误: 存在阻断级数据校验问题，禁止进入PPT阶段", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
